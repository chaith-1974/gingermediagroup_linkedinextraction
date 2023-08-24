from flask import Flask, render_template, request, send_file
import openpyxl
from googleapiclient.discovery import build
import os
import time

app = Flask(__name__)

# Initialize the Google Custom Search service
api_key = "AIzaSyA6O6ftYQ5njY1l72WQYADxsFQVa67Mi9w"

cse_id = "44bddbdbe9e9f41a3"
service = build("customsearch", "v1", developerKey=api_key)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    input_file = request.files['input_file']
    input_path = os.path.join('uploads', input_file.filename)
    input_file.save(input_path)

    output_path = process_excel(input_path)

    return send_file(output_path, as_attachment=True, download_name='output.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def process_excel(input_path):
    source_workbook = openpyxl.load_workbook(input_path)
    source_sheet = source_workbook.active

    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active

    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        keyword = row[0]
        search_query = f'{keyword} "Marketing head" site:linkedin.com/in location:"India"'

        # Perform Google Custom Search API request
        result = service.cse().list(q=search_query, cx=cse_id, num=5).execute()
        links = [item['link'] for item in result.get('items', [])]

        # Write search results to output sheet
        for link in links:
            output_sheet.append([keyword, link])

        time.sleep(5)

    output_path = 'output.xlsx'
    output_workbook.save(output_path)

    source_workbook.close()
    output_workbook.close()

    return output_path

if __name__ == '__main__':
    app.run(debug=True)
